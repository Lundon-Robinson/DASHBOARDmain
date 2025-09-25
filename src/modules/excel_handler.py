"""
Excel Handler Module
====================

Advanced Excel processing with auto-sync, duplicate detection,
pivot analysis, and report generation.
"""

import os
import shutil
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime, timedelta
import hashlib

import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import xlsxwriter

from ..core.logger import logger
from ..core.database import get_db_manager, Cardholder, Transaction

class ExcelHandler:
    """Advanced Excel processing handler"""
    
    def __init__(self, config):
        self.config = config
        self.db_manager = get_db_manager(config.database.url)
        
        # Create necessary directories
        self.templates_dir = config.paths.templates_dir
        self.exports_dir = config.paths.exports_dir
        self.temp_dir = config.paths.temp_dir
        
        # Cache for loaded files
        self._file_cache = {}
        self._last_modified = {}
        
        logger.info("Excel handler initialized")
    
    def load_treasury_data(self, file_path: str, force_reload: bool = False) -> pd.DataFrame:
        """Load treasury data from Excel file with caching and error handling"""
        try:
            file_path = Path(file_path)
            
            # Check cache
            if not force_reload and str(file_path) in self._file_cache:
                mod_time = file_path.stat().st_mtime
                if mod_time == self._last_modified.get(str(file_path)):
                    logger.debug(f"Using cached data for {file_path}")
                    return self._file_cache[str(file_path)].copy()
            
            logger.info(f"Loading treasury data from: {file_path}")
            
            # Determine file type and read accordingly
            if file_path.suffix.lower() == '.csv':
                df = pd.read_csv(file_path)
            elif file_path.suffix.lower() in ['.xls', '.xlsx', '.xlsm']:
                # Try to detect header location
                df = self._smart_read_excel(file_path)
            else:
                raise ValueError(f"Unsupported file type: {file_path.suffix}")
            
            # Clean and validate data
            df = self._clean_treasury_data(df)
            
            # Cache the data
            self._file_cache[str(file_path)] = df.copy()
            self._last_modified[str(file_path)] = file_path.stat().st_mtime
            
            logger.info(f"Successfully loaded {len(df)} treasury records")
            return df
            
        except Exception as e:
            logger.error(f"Failed to load treasury data from {file_path}", exception=e)
            raise
    
    def _smart_read_excel(self, file_path: Path) -> pd.DataFrame:
        """Intelligently read Excel file by detecting header location"""
        try:
            # First, try to find the header row by looking for common patterns
            sample_rows = pd.read_excel(file_path, nrows=10, header=None)
            
            header_row = 0
            for idx, row in sample_rows.iterrows():
                row_str = ' '.join(str(val).lower() for val in row if pd.notna(val))
                if any(keyword in row_str for keyword in ['transaction', 'amount', 'date', 'merchant', 'card']):
                    header_row = idx
                    break
            
            logger.debug(f"Detected header row at index {header_row}")
            
            # Read with detected header
            if header_row == 0:
                df = pd.read_excel(file_path)
            else:
                # Skip rows before header, use header_row as header
                df = pd.read_excel(file_path, skiprows=header_row)
            
            return df
            
        except Exception as e:
            logger.warning(f"Smart Excel reading failed, using default: {e}")
            return pd.read_excel(file_path)
    
    def _clean_treasury_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean and validate treasury data"""
        original_count = len(df)
        
        # Standardize column names (handle common variations)
        column_mapping = {
            'FIN.TRANSACTION AMOUNT': 'amount',
            'TRANSACTION AMOUNT': 'amount',
            'Amount': 'amount',
            'AMOUNT': 'amount',
            'Transaction Date': 'date',
            'DATE': 'date',
            'Date': 'date',
            'Merchant': 'merchant',
            'MERCHANT': 'merchant',
            'Card Number': 'card_number',
            'CARD NUMBER': 'card_number',
            'Card Holder': 'cardholder',
            'CARDHOLDER': 'cardholder',
        }
        
        # Rename columns
        for old_name, new_name in column_mapping.items():
            if old_name in df.columns:
                df = df.rename(columns={old_name: new_name})
        
        # Clean amount column if exists
        if 'amount' in df.columns:
            df['amount'] = pd.to_numeric(df['amount'], errors='coerce')
            # Remove rows with invalid amounts
            df = df.dropna(subset=['amount'])
        
        # Clean date column if exists
        if 'date' in df.columns:
            df['date'] = pd.to_datetime(df['date'], errors='coerce')
            df = df.dropna(subset=['date'])
        
        # Remove completely empty rows
        df = df.dropna(how='all')
        
        cleaned_count = len(df)
        if cleaned_count != original_count:
            logger.info(f"Cleaned data: {original_count} -> {cleaned_count} rows")
        
        return df
    
    def load_cardholder_data(self, file_path: str, sheet_name: str = None) -> pd.DataFrame:
        """Load cardholder information from Excel file, specifically handling OUTSTANDING LOGS sheet"""
        try:
            file_path = Path(file_path)
            logger.info(f"Loading cardholder data from: {file_path}")
            
            # Default to OUTSTANDING LOGS sheet if not specified
            if sheet_name is None:
                sheet_name = "OUTSTANDING LOGS"
            
            # Check if the sheet exists
            try:
                excel_file = pd.ExcelFile(file_path)
                available_sheets = excel_file.sheet_names
                
                if sheet_name not in available_sheets:
                    logger.warning(f"Sheet '{sheet_name}' not found. Available sheets: {available_sheets}")
                    # Try to find a similar sheet
                    for sheet in available_sheets:
                        if 'outstanding' in sheet.lower() or 'logs' in sheet.lower():
                            sheet_name = sheet
                            logger.info(f"Using sheet '{sheet_name}' instead")
                            break
                    else:
                        # Use the first sheet as fallback
                        sheet_name = available_sheets[0]
                        logger.info(f"Using first sheet '{sheet_name}' as fallback")
                
                # Read the Excel file
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                
            except Exception as e:
                logger.warning(f"Failed to read sheet '{sheet_name}': {e}")
                # Try reading without specifying sheet name
                df = pd.read_excel(file_path)
            
            # Clean and standardize cardholder data
            df = self._clean_cardholder_data(df)
            
            logger.info(f"Successfully loaded {len(df)} cardholder records from '{sheet_name}' sheet")
            return df
            
        except Exception as e:
            logger.error(f"Failed to load cardholder data from {file_path}", exception=e)
            raise
    
    def _clean_cardholder_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean and standardize cardholder data, handling OUTSTANDING LOGS sheet format"""
        
        logger.info(f"Cleaning cardholder data with columns: {list(df.columns)}")
        
        # Handle OUTSTANDING LOGS sheet format
        # Based on Bulk Mail.py, column 8 (H) contains cardholder & manager emails (semicolons)
        # Common patterns from the original script suggest positional mapping
        
        # Create a cleaned dataframe
        cleaned_df = df.copy()
        
        # Handle positional column mapping (like in Bulk Mail.py)
        if len(df.columns) >= 8:
            # Try to identify columns by position (0-indexed)
            # Column 4 and 5 often contain first and last names
            if len(df.columns) > 5:
                try:
                    first_name_col = df.iloc[:, 4] if len(df.columns) > 4 else pd.Series()
                    last_name_col = df.iloc[:, 5] if len(df.columns) > 5 else pd.Series()
                    
                    # Create full name with proper NaN handling
                    if not first_name_col.empty and not last_name_col.empty:
                        first_clean = first_name_col.fillna('').astype(str).str.replace('-', ' ').str.strip()
                        last_clean = last_name_col.fillna('').astype(str).str.strip()
                        cleaned_df['FullName'] = (first_clean + " " + last_clean).str.strip()
                except Exception as e:
                    logger.warning(f"Failed to create FullName from position: {e}")
            
            # Column 8 (index 7) typically contains email information
            if len(df.columns) > 7:
                try:
                    email_col = df.iloc[:, 7]  # Column H (8th column)
                    
                    # Handle NaN/None values first, then convert to string for str operations
                    email_col_clean = email_col.fillna('').astype(str)
                    cleaned_df['email_data'] = email_col_clean
                    
                    # Split email data on semicolons to get cardholder and manager emails
                    # Use safe string operations with proper NaN handling
                    email_split = email_col_clean.str.split(';', expand=True)
                    
                    # Extract cardholder email (first part)
                    if len(email_split.columns) >= 1:
                        cleaned_df['cardholder_email'] = email_split[0].fillna('').str.strip()
                    else:
                        cleaned_df['cardholder_email'] = ''
                    
                    # Extract manager email (second part) 
                    if len(email_split.columns) >= 2:
                        cleaned_df['manager_email'] = email_split[1].fillna('').str.strip()
                    else:
                        cleaned_df['manager_email'] = ''
                        
                except Exception as e:
                    logger.warning(f"Failed to process email column: {e}")
                    # Set default empty values if processing fails
                    cleaned_df['cardholder_email'] = ''
                    cleaned_df['manager_email'] = ''
        
        # Handle named columns if they exist
        if 'First Name' in df.columns and 'Last Name' in df.columns:
            first_clean = df['First Name'].fillna('').astype(str).str.strip()
            last_clean = df['Last Name'].fillna('').astype(str).str.strip()
            cleaned_df['FullName'] = (first_clean + " " + last_clean).str.strip()
        
        # Map common column names to standard names
        column_mapping = {
            'Section': 'department',
            'Department': 'department', 
            'Monthly Limit': 'monthly_limit',
            'Cost Centre': 'cost_centre',
            'CostCentre': 'cost_centre',
            'Email': 'email',
            'Manager Email': 'manager_email',
            'Card Number': 'card_number',
            'CardNumber': 'card_number',
            'FullName': 'name',
            'cardholder_email': 'email',
            'Full Name': 'name',
            'Cardholder Name': 'name'
        }
        
        # Rename columns if they exist
        renamed_df = cleaned_df.rename(columns=column_mapping)
        
        # Ensure we have required columns with default values
        required_columns = {
            'name': '',
            'email': '',
            'card_number': '',
            'department': '',
            'cost_centre': '',
            'manager_email': '',
            'monthly_limit': 0,
            'active': True
        }
        
        for col, default_val in required_columns.items():
            if col not in renamed_df.columns:
                renamed_df[col] = default_val
        
        # Clean up email addresses - remove NaN and invalid entries
        if 'email' in renamed_df.columns:
            # Handle NaN values first, then process as strings
            renamed_df['email'] = renamed_df['email'].fillna('').astype(str).str.strip()
            renamed_df['email'] = renamed_df['email'].replace(['nan', 'None', ''], '')
            # Basic email validation - only keep emails with @ symbol
            email_mask = renamed_df['email'].str.contains('@', na=False) & (renamed_df['email'] != '')
            renamed_df.loc[~email_mask, 'email'] = ''
        
        if 'manager_email' in renamed_df.columns:
            # Handle NaN values first, then process as strings
            renamed_df['manager_email'] = renamed_df['manager_email'].fillna('').astype(str).str.strip()
            renamed_df['manager_email'] = renamed_df['manager_email'].replace(['nan', 'None', ''], '')
            # Basic email validation - only keep emails with @ symbol
            manager_email_mask = renamed_df['manager_email'].str.contains('@', na=False) & (renamed_df['manager_email'] != '')
            renamed_df.loc[~manager_email_mask, 'manager_email'] = ''
        
        # Clean up names
        if 'name' in renamed_df.columns:
            renamed_df['name'] = renamed_df['name'].fillna('').astype(str).str.strip()
            renamed_df['name'] = renamed_df['name'].replace(['nan', 'None', ''], '')
        
        # Remove rows with no name OR email (more lenient filtering)
        # At least one of name or email should be present and valid
        before_count = len(renamed_df)
        renamed_df = renamed_df[
            ((renamed_df['name'].notna()) & (renamed_df['name'] != '')) | 
            ((renamed_df['email'].notna()) & (renamed_df['email'] != '') & (renamed_df['email'].str.contains('@', na=False)))
        ]
        after_count = len(renamed_df)
        
        if before_count != after_count:
            logger.info(f"Filtered out {before_count - after_count} incomplete records")
        
        # Add timestamps
        from datetime import datetime
        now = datetime.now()
        renamed_df['created_at'] = now
        renamed_df['updated_at'] = now
        
        logger.info(f"Cleaned cardholder data: {len(renamed_df)} valid records")
        return renamed_df
        
        # Handle positional columns if named columns don't exist
        if len(df.columns) >= 14:
            if 'department' not in df.columns and len(df.columns) > 6:
                df['department'] = df.iloc[:, 6]
            if 'monthly_limit' not in df.columns and len(df.columns) > 12:
                df['monthly_limit'] = df.iloc[:, 12]
            if 'cost_centre' not in df.columns and len(df.columns) > 13:
                df['cost_centre'] = df.iloc[:, 13]
        
        # Ensure required fields exist
        required_fields = ['FullName', 'email', 'card_number']
        for field in required_fields:
            if field not in df.columns:
                if field == 'FullName' and 'Name' in df.columns:
                    df['FullName'] = df['Name']
                elif field not in df.columns:
                    logger.warning(f"Required field '{field}' not found in data")
        
        # Clean up the data
        df = df.dropna(how='all')
        
        # Remove rows where essential data is missing
        if 'FullName' in df.columns:
            df = df.dropna(subset=['FullName'])
        
        logger.info(f"Cleaned cardholder data: {len(df)} records with columns {list(df.columns)}")
        return df
    
    def detect_duplicates(self, df: pd.DataFrame, key_columns: List[str] = None) -> pd.DataFrame:
        """Detect duplicate records in DataFrame"""
        if key_columns is None:
            key_columns = ['date', 'amount', 'merchant'] if all(col in df.columns for col in ['date', 'amount', 'merchant']) else df.columns[:3]
        
        # Find duplicates
        duplicates = df.duplicated(subset=key_columns, keep=False)
        duplicate_df = df[duplicates].copy()
        
        if len(duplicate_df) > 0:
            logger.warning(f"Found {len(duplicate_df)} duplicate records")
        
        return duplicate_df
    
    def merge_data_smart(self, df1: pd.DataFrame, df2: pd.DataFrame, 
                        merge_strategy: str = "outer") -> pd.DataFrame:
        """Intelligently merge two DataFrames"""
        try:
            # Find common columns for merging
            common_cols = set(df1.columns) & set(df2.columns)
            
            if not common_cols:
                logger.warning("No common columns found, concatenating DataFrames")
                return pd.concat([df1, df2], ignore_index=True)
            
            # Try to find the best merge key
            merge_key = None
            for col in ['id', 'card_number', 'transaction_id', 'date']:
                if col in common_cols:
                    merge_key = col
                    break
            
            if not merge_key:
                merge_key = list(common_cols)[0]
            
            logger.info(f"Merging DataFrames on column: {merge_key}")
            merged = pd.merge(df1, df2, on=merge_key, how=merge_strategy, suffixes=('', '_y'))
            
            # Remove duplicate columns from merge
            merged = merged.loc[:, ~merged.columns.str.endswith('_y')]
            
            return merged
            
        except Exception as e:
            logger.error("Smart merge failed, falling back to concatenation", exception=e)
            return pd.concat([df1, df2], ignore_index=True)
    
    def create_pivot_analysis(self, df: pd.DataFrame, index_col: str, 
                            value_col: str, agg_func: str = 'sum') -> pd.DataFrame:
        """Create pivot table analysis"""
        try:
            pivot_table = pd.pivot_table(
                df, 
                index=index_col, 
                values=value_col, 
                aggfunc=agg_func, 
                fill_value=0
            )
            
            logger.info(f"Created pivot analysis: {len(pivot_table)} groups")
            return pivot_table
            
        except Exception as e:
            logger.error("Failed to create pivot analysis", exception=e)
            raise
    
    def generate_statement(self, cardholder_id: int, period_start: datetime, 
                         period_end: datetime, template_path: str = None) -> str:
        """Generate individual statement for cardholder"""
        try:
            # Get cardholder data
            cardholder = self.db_manager.get_session().query(Cardholder).filter(
                Cardholder.id == cardholder_id
            ).first()
            
            if not cardholder:
                raise ValueError(f"Cardholder with ID {cardholder_id} not found")
            
            # Get transactions for period
            transactions = self.db_manager.get_transactions(
                cardholder_id=cardholder_id,
                start_date=period_start,
                end_date=period_end
            )
            
            # Create statement workbook
            if template_path and Path(template_path).exists():
                wb = load_workbook(template_path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Statement"
            
            # Add header information
            ws['A1'] = "Purchase Card Statement"
            ws['A2'] = f"Cardholder: {cardholder.name}"
            ws['A3'] = f"Card Number: {cardholder.card_number}"
            ws['A4'] = f"Period: {period_start.strftime('%d/%m/%Y')} - {period_end.strftime('%d/%m/%Y')}"
            
            # Add transaction headers
            headers = ['Date', 'Merchant', 'Amount', 'Currency', 'Category', 'Description']
            for col, header in enumerate(headers, 1):
                ws.cell(row=6, column=col, value=header)
            
            # Add transaction data
            total_amount = 0
            for row, transaction in enumerate(transactions, 7):
                ws.cell(row=row, column=1, value=transaction.transaction_date.strftime('%d/%m/%Y'))
                ws.cell(row=row, column=2, value=transaction.merchant)
                ws.cell(row=row, column=3, value=transaction.amount)
                ws.cell(row=row, column=4, value=transaction.currency)
                ws.cell(row=row, column=5, value=transaction.category)
                ws.cell(row=row, column=6, value=transaction.description)
                total_amount += transaction.amount
            
            # Add total
            total_row = len(transactions) + 7
            ws.cell(row=total_row, column=2, value="Total:")
            ws.cell(row=total_row, column=3, value=total_amount)
            
            # Style the statement
            self._style_statement(ws, len(transactions))
            
            # Save the statement
            statement_filename = f"Statement_{cardholder.name.replace(' ', '_')}_{period_start.strftime('%Y%m')}.xlsx"
            statement_path = self.exports_dir / statement_filename
            wb.save(statement_path)
            
            logger.info(f"Generated statement: {statement_path}")
            return str(statement_path)
            
        except Exception as e:
            logger.error("Failed to generate statement", exception=e)
            raise
    
    def _style_statement(self, ws, data_rows: int):
        """Apply styling to statement worksheet"""
        # Header styling
        header_font = Font(size=14, bold=True)
        ws['A1'].font = header_font
        
        # Column headers styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col in range(1, 7):  # A to F
            cell = ws.cell(row=6, column=col)
            cell.fill = header_fill
            cell.font = header_font
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(6, data_rows + 8):
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = thin_border
    
    def batch_generate_statements(self, period_start: datetime, period_end: datetime,
                                template_path: str = None) -> List[str]:
        """Generate statements for all active cardholders"""
        try:
            cardholders = self.db_manager.get_cardholders(active_only=True)
            statement_paths = []
            
            logger.info(f"Generating statements for {len(cardholders)} cardholders")
            
            for cardholder in cardholders:
                try:
                    path = self.generate_statement(
                        cardholder.id, 
                        period_start, 
                        period_end, 
                        template_path
                    )
                    statement_paths.append(path)
                except Exception as e:
                    logger.error(f"Failed to generate statement for {cardholder.name}", exception=e)
            
            logger.info(f"Successfully generated {len(statement_paths)} statements")
            return statement_paths
            
        except Exception as e:
            logger.error("Batch statement generation failed", exception=e)
            raise
    
    def export_to_pdf(self, excel_path: str, pdf_path: str = None) -> str:
        """Convert Excel file to PDF (requires additional dependencies)"""
        # This would require additional libraries like reportlab or win32com
        # For now, return the original path
        logger.warning("PDF export not yet implemented")
        return excel_path
    
    def export_to_html(self, df: pd.DataFrame, html_path: str = None) -> str:
        """Export DataFrame to HTML with styling"""
        if html_path is None:
            html_path = self.exports_dir / f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
        
        # Create styled HTML
        html_content = df.to_html(
            table_id="dashboard-table",
            classes="table table-striped table-bordered",
            escape=False
        )
        
        # Add CSS styling
        styled_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Dashboard Export</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                .table {{ width: 100%; border-collapse: collapse; }}
                .table th, .table td {{ padding: 8px; text-align: left; border: 1px solid #ddd; }}
                .table th {{ background-color: #f2f2f2; font-weight: bold; }}
                .table-striped tbody tr:nth-child(odd) {{ background-color: #f9f9f9; }}
                h1 {{ color: #333; }}
            </style>
        </head>
        <body>
            <h1>Dashboard Export - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</h1>
            {html_content}
        </body>
        </html>
        """
        
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(styled_html)
        
        logger.info(f"Exported to HTML: {html_path}")
        return str(html_path)
    
    def validate_data(self, df: pd.DataFrame, rules: Dict[str, Any]) -> Dict[str, List]:
        """Validate data against business rules"""
        validation_results = {
            'errors': [],
            'warnings': [],
            'info': []
        }
        
        for column, rule in rules.items():
            if column not in df.columns:
                validation_results['errors'].append(f"Column '{column}' not found")
                continue
            
            if 'required' in rule and rule['required']:
                null_count = df[column].isnull().sum()
                if null_count > 0:
                    validation_results['errors'].append(f"Column '{column}' has {null_count} null values")
            
            if 'min_value' in rule:
                invalid = df[df[column] < rule['min_value']]
                if len(invalid) > 0:
                    validation_results['warnings'].append(f"Column '{column}' has {len(invalid)} values below minimum {rule['min_value']}")
            
            if 'max_value' in rule:
                invalid = df[df[column] > rule['max_value']]
                if len(invalid) > 0:
                    validation_results['warnings'].append(f"Column '{column}' has {len(invalid)} values above maximum {rule['max_value']}")
        
        return validation_results
    
    def sync_with_database(self, df: pd.DataFrame, table_type: str = 'transactions'):
        """Sync DataFrame data with database"""
        try:
            if table_type == 'transactions':
                self._sync_transactions(df)
            elif table_type == 'cardholders':
                self._sync_cardholders(df)
            else:
                raise ValueError(f"Unknown table type: {table_type}")
                
        except Exception as e:
            logger.error(f"Database sync failed for {table_type}", exception=e)
            raise
    
    def _sync_transactions(self, df: pd.DataFrame):
        """Sync transaction data with database"""
        logger.info(f"Syncing {len(df)} transactions with database")
        
        synced_count = 0
        for _, row in df.iterrows():
            try:
                # Find cardholder by card number or name
                cardholder = None
                if 'card_number' in row and pd.notna(row['card_number']):
                    cardholder = self.db_manager.get_cardholder_by_card_number(str(row['card_number']))
                
                if not cardholder and 'cardholder' in row and pd.notna(row['cardholder']):
                    # Try to find by name
                    with self.db_manager.get_session() as session:
                        cardholder = session.query(Cardholder).filter(
                            Cardholder.name.ilike(f"%{row['cardholder']}%")
                        ).first()
                
                if not cardholder:
                    logger.warning(f"Cardholder not found for transaction: {row.get('merchant', 'Unknown')}")
                    continue
                
                # Create transaction
                transaction = self.db_manager.create_transaction(
                    cardholder_id=cardholder.id,
                    transaction_date=pd.to_datetime(row.get('date', datetime.now())),
                    merchant=str(row.get('merchant', '')),
                    amount=float(row.get('amount', 0)),
                    currency=str(row.get('currency', 'GBP')),
                    category=str(row.get('category', '')),
                    description=str(row.get('description', ''))
                )
                synced_count += 1
                
            except Exception as e:
                logger.warning(f"Failed to sync transaction: {e}")
        
        logger.info(f"Successfully synced {synced_count} transactions")
    
    def _sync_cardholders(self, df: pd.DataFrame):
        """Sync cardholder data with database"""
        logger.info(f"Syncing {len(df)} cardholders with database")
        
        synced_count = 0
        for _, row in df.iterrows():
            try:
                # Check if cardholder already exists
                existing = None
                if 'card_number' in row and pd.notna(row['card_number']):
                    existing = self.db_manager.get_cardholder_by_card_number(str(row['card_number']))
                
                if existing:
                    # Update existing cardholder
                    with self.db_manager.get_session() as session:
                        existing.name = str(row.get('name', existing.name))  # Use 'name' not 'FullName'
                        existing.email = str(row.get('email', existing.email))
                        existing.department = str(row.get('department', existing.department) if pd.notna(row.get('department')) else existing.department)
                        existing.manager_email = str(row.get('manager_email', existing.manager_email) if pd.notna(row.get('manager_email')) else existing.manager_email)
                        existing.cost_centre = str(row.get('cost_centre', existing.cost_centre) if pd.notna(row.get('cost_centre')) else existing.cost_centre)
                        existing.updated_at = datetime.utcnow()
                        session.commit()
                else:
                    # Create new cardholder
                    self.db_manager.create_cardholder(
                        card_number=str(row.get('card_number', f"TEMP_{synced_count}")),
                        name=str(row.get('name', 'Unknown')),  # Use 'name' not 'FullName'
                        email=str(row.get('email', '')),
                        manager_email=str(row.get('manager_email', '') if pd.notna(row.get('manager_email')) else None),
                        department=str(row.get('department', '') if pd.notna(row.get('department')) else None),
                        cost_centre=str(row.get('cost_centre', '') if pd.notna(row.get('cost_centre')) else None)
                    )
                
                synced_count += 1
                
            except Exception as e:
                logger.warning(f"Failed to sync cardholder: {e}")
        
        logger.info(f"Successfully synced {synced_count} cardholders")

__all__ = ['ExcelHandler']