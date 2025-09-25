import os
import traceback
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk

# Local log file for user-facing diagnostics
LOG_FILE = r"C:\Users\NADLUROB\Desktop\Dash\log.txt"


def _log_and_alert(exc: Exception, context: str = ""):
    """Log exception with traceback to LOG_FILE and show a messagebox with actionable text.

    This helper centralizes user-facing error messages so the GUI can present
    concise, helpful advice while the full traceback is written to the log file.
    """
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(f"\n[{datetime.now():%Y-%m-%d %H:%M:%S}] ERROR in {context}\n")
            f.write(''.join(traceback.format_exception(type(exc), exc, exc.__traceback__)))
            f.write("\n" + "-" * 60 + "\n")
    except Exception:
        # best-effort logging; if it fails we still proceed to show a messagebox
        pass

    # Show a concise message to the user with a suggestion for remediation
    try:
        hint = "Check file paths, Excel file formats (.xls/.xlsx/.xlsm), and that required applications like Excel are installed."
        full_msg = f"{str(exc)}\n\nContext: {context}\n\nSuggested action: {hint}"
        # Use tkinter messagebox if available
        try:
            messagebox.showerror("Error", full_msg)
        except Exception:
            print("ERROR:", full_msg)
    except Exception:
        # If all UI notification fails, print to stdout
        print("An error occurred:", exc)

# --- Your existing functions (unchanged) ---
def load_treasury_data(treasury_file):
    """Load treasury data from the given Excel file and return a cleaned DataFrame."""
    try:
        print(f"Loading treasury data from: {treasury_file}")
        headers_df = pd.read_excel(treasury_file, header=None, nrows=1, skiprows=1)
        headers = headers_df.iloc[0].tolist()
        print(f"Headers found: {headers}")
    except Exception as e:
        _log_and_alert(e, context=f"Reading treasury headers: {treasury_file}")
        raise

    try:
        data_df = pd.read_excel(treasury_file, header=None, skiprows=3)
    except Exception as e:
        _log_and_alert(e, context=f"Reading treasury data: {treasury_file}")
        raise
    data_df.columns = headers

    before_drop = len(data_df)
    data_df = data_df[pd.to_numeric(data_df['FIN.TRANSACTION AMOUNT'], errors='coerce').notnull()]
    after_drop = len(data_df)
    print(f"Dropped {before_drop - after_drop} non-numeric FIN.TRANSACTION AMOUNT rows")
    print(f"Loaded treasury data rows: {len(data_df)}")

    return data_df

def load_cardholder_info(lookup_path):
    """Load cardholder information from the lookup Excel file and return a DataFrame."""
    try:
        df = pd.read_excel(lookup_path, sheet_name="Sheet1", header=None)
    except Exception as e:
        _log_and_alert(e, context=f"Reading cardholder lookup file: {lookup_path}")
        raise

    df['FullName'] = df.iloc[:, 4].astype(str).str.replace('-', ' ', regex=False).str.strip() + " " + df.iloc[:, 5].astype(str).str.strip()
    df['Section'] = df.iloc[:, 6]
    df['MonthlyLimit'] = df.iloc[:, 12]
    df['CostCentre'] = df.iloc[:, 13]

    df = df[['FullName', 'Section', 'MonthlyLimit', 'CostCentre']]
    return df

import xlwings as xw

def clear_and_copy_to_template_xw(working_template_file, treasury_data_df, data_tab_name='Data'):
    """Clear the data tab in the template and copy treasury data using xlwings."""
    print(f"Opening working template file with xlwings: {working_template_file} and clearing '{data_tab_name}' tab")
    app = None
    wb = None
    try:
        app = xw.App(visible=False)
        wb = app.books.open(working_template_file)
        if data_tab_name not in [sheet.name for sheet in wb.sheets]:
            raise Exception(f"Data tab '{data_tab_name}' not found in template.")

        ws = wb.sheets[data_tab_name]
        ws.clear_contents()
        ws.range("A1").value = list(treasury_data_df.columns)
        ws.range("A2").value = treasury_data_df.values.tolist()
        wb.save()
        print(f"Copied treasury data to '{data_tab_name}' tab and saved working template.")
    except Exception as e:
        _log_and_alert(e, context=f"Updating template: {working_template_file}")
        raise
    finally:
        try:
            if wb:
                wb.close()
        except Exception:
            pass
        try:
            if app:
                app.quit()
        except Exception:
            pass


def read_cardfiles_from_create_files(template_path, create_files_tab='Create Files'):
    """Read cardfile rows from the specified tab in the template and return a list of tuples."""
    print(f"Reading cardfile rows from '{create_files_tab}' tab in {template_path}")
    cardfiles = []
    app = None
    wb = None
    try:
        app = xw.App(visible=False)
        wb = app.books.open(template_path)
        ws = wb.sheets[create_files_tab]

        last_row = ws.range("C" + str(ws.cells.last_cell.row)).end('up').row
        for row in range(3, last_row + 1):
            filename = ws.range(f"C{row}").value
            card_number = ws.range(f"D{row}").value
            if all([card_number, filename]):
                clean_filename = str(filename).replace(" for FIN.TRANSACTION DATE", "").strip()
                if '\\' in clean_filename:
                    clean_filename = clean_filename.split('\\')[-1]
                if '/' in clean_filename:
                    clean_filename = clean_filename.split('/')[-1]
                cardfiles.append((str(card_number).strip(), clean_filename))
        print(f"Read {len(cardfiles)} cardfiles entries")
    except Exception as e:
        _log_and_alert(e, context=f"Reading create files tab: {template_path}")
        raise
    finally:
        try:
            if wb:
                wb.close()
        except Exception:
            pass
        try:
            if app:
                app.quit()
        except Exception:
            pass
    return cardfiles


def create_individual_statements(
    working_template_file,
    base_blank_file,
    cardfiles,
    statements_folder,
    cardholder_lookup_df,
    filename_suffix="",  # Added parameter for filename suffix
    max_warnings_per_type=10
):
    """Create individual statements for each cardholder and save them as Excel files."""

    try:
        print(f"Creating individual statements into folder: {statements_folder}")
        os.makedirs(statements_folder, exist_ok=True)

        wb = load_workbook(working_template_file, data_only=True)
        if 'Data' not in wb.sheetnames:
            raise Exception("Template does not contain 'Data' sheet.")
        data_ws = wb['Data']
    except Exception as e:
        _log_and_alert(e, context=f"Preparing statements from template: {working_template_file}")
        raise

    # Read data sheet to DataFrame
    data_rows = []
    headers = []
    for i, row in enumerate(data_ws.iter_rows(values_only=True)):
        if i == 0:
            headers = list(row)
        else:
            data_rows.append(row)
    df_data = pd.DataFrame(data_rows, columns=headers)

    warning_counts = {
        "invalid_filename": 0,
        "no_data_found": 0
    }
    created_count = 0

    for card_number, filename in cardfiles:
        try:
            if not filename:
                if warning_counts["invalid_filename"] < max_warnings_per_type:
                    print(f"[WARNING] Skipping empty filename for card number: {card_number}")
                warning_counts["invalid_filename"] += 1
                continue

            filtered_df = df_data[df_data['ACC.ACCOUNT NUMBER'] == card_number]
            if filtered_df.empty:
                if warning_counts["no_data_found"] < max_warnings_per_type:
                    print(f"[WARNING] No data for card number: {card_number}")
                warning_counts["no_data_found"] += 1
                continue

            full_name = filtered_df.iloc[0]['ACC.ACCOUNT NAME']

            try:
                statement_wb = load_workbook(base_blank_file)
            except Exception as e:
                _log_and_alert(e, context=f"Opening blank statement file: {base_blank_file}")
                warning_counts.setdefault("template_open_fail", 0)
                warning_counts["template_open_fail"] += 1
                continue

            if 'Statement' not in statement_wb.sheetnames:
                print(f"[WARNING] 'Statement' sheet missing in template: {base_blank_file}")
                try:
                    statement_wb.close()
                except Exception:
                    pass
                continue
            sheet = statement_wb['Statement']

            # Clear rows A12:F
            for row in sheet.iter_rows(min_row=12, max_row=307, min_col=1, max_col=6):
                for cell in row:
                    cell.value = None

            # Populate header
            sheet['B3'] = full_name
            sheet['B5'] = card_number
            sheet['B7'] = datetime.now().strftime('%d-%b-%Y')

            # Lookup extras: Monthly Limit (M), Section (G), Cost Centre (N)
            lookup = cardholder_lookup_df[cardholder_lookup_df['FullName'].str.lower() == full_name.strip().lower()]
            if not lookup.empty:
                sheet['F7'] = lookup.iloc[0]['CostCentre']
                sheet['F5'] = lookup.iloc[0]['Section']
                sheet['B8'] = lookup.iloc[0]['MonthlyLimit']

            # Write transactions starting A12
            for idx, row in enumerate(filtered_df.itertuples(index=False), start=12):
                sheet.cell(row=idx, column=1, value=row[2])  # Transaction Date (col C)
                sheet.cell(row=idx, column=2, value=row[3])  # Post Date (col D)
                sheet.cell(row=idx, column=3, value=row[14])  # Bank Reference (col L)
                sheet.cell(row=idx, column=4, value=row[9])  # Your Reference (col J)
                sheet.cell(row=idx, column=5, value=row[8])  # Description (col I)

                value = row[12]  # Transaction Value (col M)
                cell = sheet.cell(row=idx, column=6, value=value)
                if isinstance(value, (int, float)):
                    cell.number_format = u'£#,##0.00'
            from openpyxl.styles import Alignment

            # Calculate and set sum to F8
            last_row = 11 + len(filtered_df)
            total_cell = sheet['F8']
            total_cell.value = f"=SUM(F12:F{last_row})"
            total_cell.number_format = u'£#,##0.00'

            alignment_left = Alignment(horizontal='left')
            for row in sheet.iter_rows(min_row=12, max_row=last_row, min_col=1, max_col=6):
                for cell in row:
                    val = cell.value
                    if isinstance(val, str):
                        val = val.strip()
                        if val:
                            cell.value = val.title()

                    cell.alignment = alignment_left

            # Format B3, F5, B5, etc. to Proper + left-aligned if text
            for cell in [sheet['B3'], sheet['B5'], sheet['B7'], sheet['F5'], sheet['F7']]:
                val = cell.value
                if isinstance(val, str):
                    val = val.strip()
                    cell.value = val.title()
                    cell.alignment = alignment_left

            # Save as .xlsx with suffix replacing old fixed suffix
            if not filename.lower().endswith(".xlsx"):
                filename += ".xlsx"

            # Insert the filename_suffix before .xlsx extension
            name_only, ext = os.path.splitext(filename)
            filename_with_suffix = f"{name_only} {filename_suffix}{ext}"

            save_path = os.path.join(statements_folder, filename_with_suffix)
            try:
                statement_wb.save(save_path)
                created_count += 1
                print(f"Saved: {save_path}")
            except Exception as e:
                _log_and_alert(e, context=f"Saving statement for card {card_number} to {save_path}")
                warning_counts.setdefault("save_failures", 0)
                warning_counts["save_failures"] += 1
            finally:
                try:
                    statement_wb.close()
                except Exception:
                    pass

        except Exception as e:
            _log_and_alert(e, context=f"Processing card {card_number}")
            warning_counts.setdefault("processing_errors", 0)
            warning_counts["processing_errors"] += 1
            continue

    # Final warnings
    for k, v in warning_counts.items():
        if v > max_warnings_per_type:
            print(f"[NOTE] Suppressed {v - max_warnings_per_type} additional '{k}' warnings.")

    print(f"\nTotal statements created: {created_count}")


# --- GUI & main logic integration ---

DEFAULTS = {
    "treasury_file": r"C:\Users\NADLUROB\Desktop\test\MB_DSC_Monthly_Report.xls",
    "process_template_file": r"C:\Users\NADLUROB\Desktop\test\NEW Process Template DSC 2025.xlsm",
    "base_blank_file": r"C:\Users\NADLUROB\Desktop\test\Blank - Active from September 2024 .xlsx",
    "output_folder": r"C:\Users\NADLUROB\Desktop\test\output",
    "cardholder_list_file": r"\\reiltys\iomgroot\DeptShare_DHSS_Nobles\Management\Director of Finance, Performance & Delivery\16. Manx Care\FAS DSC\Purchase Cards info\Card Holder List\Purchase cardholder list DSC.xls",
}

def browse_file(entry_widget):
    filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xls *.xlsx *.xlsm")])
    if filepath:
        entry_widget.delete(0, tk.END)
        entry_widget.insert(0, filepath)

def browse_folder(entry_widget):
    folderpath = filedialog.askdirectory()
    if folderpath:
        entry_widget.delete(0, tk.END)
        entry_widget.insert(0, folderpath)

def run_gui():
    root = tk.Tk()
    root.title("DSC Statement Generator")

    frame = ttk.Frame(root, padding=10)
    frame.grid(row=0, column=0, sticky="nsew")

    # Treasury File
    ttk.Label(frame, text="Treasury Report File:").grid(row=0, column=0, sticky="w")
    treasury_entry = ttk.Entry(frame, width=70)
    treasury_entry.grid(row=0, column=1, sticky="w")
    treasury_entry.insert(0, DEFAULTS["treasury_file"])
    ttk.Button(frame, text="Browse", command=lambda: browse_file(treasury_entry)).grid(row=0, column=2, padx=5)

    # Process Template File
    ttk.Label(frame, text="Process Template File:").grid(row=1, column=0, sticky="w")
    process_entry = ttk.Entry(frame, width=70)
    process_entry.grid(row=1, column=1, sticky="w")
    process_entry.insert(0, DEFAULTS["process_template_file"])
    ttk.Button(frame, text="Browse", command=lambda: browse_file(process_entry)).grid(row=1, column=2, padx=5)

    # Base Blank File
    ttk.Label(frame, text="Base Blank Statement File:").grid(row=2, column=0, sticky="w")
    blank_entry = ttk.Entry(frame, width=70)
    blank_entry.grid(row=2, column=1, sticky="w")
    blank_entry.insert(0, DEFAULTS["base_blank_file"])
    ttk.Button(frame, text="Browse", command=lambda: browse_file(blank_entry)).grid(row=2, column=2, padx=5)

    # Output Folder
    ttk.Label(frame, text="Output Folder:").grid(row=3, column=0, sticky="w")
    output_entry = ttk.Entry(frame, width=70)
    output_entry.grid(row=3, column=1, sticky="w")
    output_entry.insert(0, DEFAULTS["output_folder"])
    ttk.Button(frame, text="Browse", command=lambda: browse_folder(output_entry)).grid(row=3, column=2, padx=5)

    # Month and Year selectors
    ttk.Label(frame, text="Select Month:").grid(row=4, column=0, sticky="w")
    month_var = tk.StringVar(value="Jul")
    month_combobox = ttk.Combobox(frame, textvariable=month_var, values=[
        "Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"], state="readonly", width=5)
    month_combobox.grid(row=4, column=1, sticky="w")

    ttk.Label(frame, text="Select Year:").grid(row=5, column=0, sticky="w")
    year_var = tk.StringVar(value="25")
    year_combobox = ttk.Combobox(frame, textvariable=year_var, values=[str(y) for y in range(20, 40)], state="readonly", width=5)
    year_combobox.grid(row=5, column=1, sticky="w")

    def on_run():
        treasury_file = treasury_entry.get().strip()
        process_template_file = process_entry.get().strip()
        base_blank_file = blank_entry.get().strip()
        output_folder = output_entry.get().strip()
        month = month_var.get()
        year = year_var.get()

        # Validation
        missing = []
        if not os.path.isfile(treasury_file): missing.append("Treasury Report File")
        if not os.path.isfile(process_template_file): missing.append("Process Template File")
        if not os.path.isfile(base_blank_file): missing.append("Base Blank Statement File")
        if not os.path.isdir(output_folder): missing.append("Output Folder")

        if missing:
            messagebox.showerror("Error", f"Please select valid files/folder for:\n" + "\n".join(missing))
            return

        try:
            main(
                treasury_file,
                process_template_file,
                base_blank_file,
                output_folder,
                month,
                year
            )
            messagebox.showinfo("Success", "Statements created successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")

    ttk.Button(frame, text="Run", command=on_run).grid(row=6, column=1, pady=10)

    root.mainloop()


def main(
    treasury_file,
    process_template_file,
    base_blank_file,
    output_folder,
    month,
    year
):
    print("Starting process...")

    treasury_data_df = load_treasury_data(treasury_file)
    cardholder_lookup_df = load_cardholder_info(DEFAULTS["cardholder_list_file"])

    # Use xlwings version to update Data tab safely in .xlsm file
    clear_and_copy_to_template_xw(process_template_file, treasury_data_df, data_tab_name='Data')

    cardfiles = read_cardfiles_from_create_files(process_template_file, create_files_tab='Create Files')

    filename_suffix = f"for {month} {year}"

    create_individual_statements(
        working_template_file=process_template_file,
        base_blank_file=base_blank_file,
        cardfiles=cardfiles,
        statements_folder=output_folder,
        cardholder_lookup_df=cardholder_lookup_df,
        filename_suffix=filename_suffix
    )


if __name__ == "__main__":
    run_gui()
